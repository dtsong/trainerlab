"""Data export service for creator CSV/JSON/XLSX exports."""

import asyncio
import csv
import json
import logging
from datetime import UTC, datetime, timedelta
from io import BytesIO, StringIO
from typing import Any
from uuid import UUID, uuid4

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.models.data_export import DataExport
from src.models.meta_snapshot import MetaSnapshot
from src.models.tournament import Tournament
from src.models.user import User
from src.services.storage_service import StorageService

logger = logging.getLogger(__name__)


EXPORT_TYPES = {
    "meta_snapshot",
    "meta_history",
    "tournament_results",
    "archetype_evolution",
    "card_usage",
    "jp_data",
}

EXPORT_FORMATS = {"csv", "json", "xlsx"}

CONTENT_TYPES = {
    "csv": "text/csv",
    "json": "application/json",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


class DataExportService:
    """Service for creating data exports."""

    def __init__(
        self,
        session: AsyncSession,
        storage: StorageService | None = None,
    ) -> None:
        self.session = session
        self.storage = storage or StorageService()

    async def create_export(
        self,
        user: User,
        export_type: str,
        config: dict[str, Any],
        format: str = "json",
    ) -> DataExport:
        """Create a new data export.

        Args:
            user: The authenticated creator user
            export_type: Type of export (meta_snapshot, etc.)
            config: Export configuration
            format: Output format (csv, json, xlsx)

        Returns:
            DataExport model with file URL

        Raises:
            ValueError: If export type or format is invalid
        """
        if export_type not in EXPORT_TYPES:
            raise ValueError(f"Invalid export type: {export_type}")
        if format not in EXPORT_FORMATS:
            raise ValueError(f"Invalid format: {format}")

        # Create pending export record
        export = DataExport(
            id=uuid4(),
            user_id=user.id,
            export_type=export_type,
            config=config,
            format=format,
            status="pending",
        )
        self.session.add(export)
        await self.session.commit()
        await self.session.refresh(export)

        try:
            # Fetch data
            data = await self._fetch_export_data(export_type, config)

            # Generate formatted content
            content, columns = await self._generate_content(data, format)

            # Upload to GCS
            filename = f"{export.id}.{format}"
            content_type = CONTENT_TYPES[format]
            file_url = await self.storage.upload_export(content, filename, content_type)

            # Update export record
            export.file_path = file_url
            export.file_size_bytes = len(content)
            export.status = "completed"
            export.expires_at = datetime.now(UTC) + timedelta(hours=24)

            await self.session.commit()
            await self.session.refresh(export)

            logger.info("Created export %s for user %s", export.id, user.id)
            return export

        except Exception as e:
            export.status = "failed"
            export.error_message = str(e)
            await self.session.commit()
            logger.exception("Failed to create export %s", export.id)
            raise

    async def get_export(self, export_id: UUID, user: User) -> DataExport | None:
        """Get an export by ID.

        Args:
            export_id: Export ID
            user: The authenticated user (must own the export)

        Returns:
            DataExport if found and owned, None otherwise
        """
        query = select(DataExport).where(
            DataExport.id == export_id,
            DataExport.user_id == user.id,
        )
        result = await self.session.execute(query)
        return result.scalar_one_or_none()

    async def list_user_exports(
        self,
        user: User,
        limit: int = 20,
    ) -> list[DataExport]:
        """List exports for a user.

        Args:
            user: The authenticated user
            limit: Maximum number of exports to return

        Returns:
            List of exports
        """
        query = (
            select(DataExport)
            .where(DataExport.user_id == user.id)
            .order_by(DataExport.created_at.desc())
            .limit(limit)
        )
        result = await self.session.execute(query)
        return list(result.scalars().all())

    async def generate_download_url(
        self,
        export_id: UUID,
        user: User,
    ) -> str | None:
        """Generate a signed download URL for an export.

        Args:
            export_id: Export ID
            user: The authenticated user (must own the export)

        Returns:
            Signed URL if export exists and is completed, None otherwise
        """
        export = await self.get_export(export_id, user)
        if not export or export.status != "completed":
            return None

        if export.expires_at and export.expires_at < datetime.now(UTC):
            return None

        # Extract filename from path
        filename = f"{export.id}.{export.format}"
        return await self.storage.generate_signed_url(filename)

    async def _fetch_export_data(
        self,
        export_type: str,
        config: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """Fetch data for export based on type and config."""
        if export_type == "meta_snapshot":
            return await self._fetch_meta_snapshot(config)
        elif export_type == "meta_history":
            return await self._fetch_meta_history(config)
        elif export_type == "tournament_results":
            return await self._fetch_tournament_results(config)
        elif export_type == "jp_data":
            return await self._fetch_jp_data(config)
        else:
            raise ValueError(f"Unsupported export type: {export_type}")

    async def _fetch_meta_snapshot(
        self,
        config: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """Fetch current meta snapshot data."""
        region = config.get("region")
        format_type = config.get("format", "standard")
        best_of = config.get("best_of", 3)

        query = (
            select(MetaSnapshot)
            .where(MetaSnapshot.format == format_type)
            .where(MetaSnapshot.best_of == best_of)
        )

        if region:
            query = query.where(MetaSnapshot.region == region)
        else:
            query = query.where(MetaSnapshot.region.is_(None))

        query = query.order_by(MetaSnapshot.snapshot_date.desc()).limit(1)

        result = await self.session.execute(query)
        snapshot = result.scalar_one_or_none()

        if not snapshot:
            return []

        data = []
        for archetype, share in snapshot.archetype_shares.items():
            tier_assignments = snapshot.tier_assignments
            tier = tier_assignments.get(archetype) if tier_assignments else None
            trend = snapshot.trends.get(archetype, {}) if snapshot.trends else {}
            data.append(
                {
                    "snapshot_date": snapshot.snapshot_date.isoformat(),
                    "region": region or "Global",
                    "format": format_type,
                    "archetype": archetype,
                    "share": float(share),
                    "tier": tier,
                    "trend": trend.get("direction"),
                    "trend_change": trend.get("change"),
                }
            )

        return sorted(data, key=lambda x: x["share"], reverse=True)

    async def _fetch_meta_history(
        self,
        config: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """Fetch meta history data."""
        region = config.get("region")
        format_type = config.get("format", "standard")
        best_of = config.get("best_of", 3)
        days = config.get("days", 30)

        query = (
            select(MetaSnapshot)
            .where(MetaSnapshot.format == format_type)
            .where(MetaSnapshot.best_of == best_of)
        )

        if region:
            query = query.where(MetaSnapshot.region == region)
        else:
            query = query.where(MetaSnapshot.region.is_(None))

        query = query.order_by(MetaSnapshot.snapshot_date.desc()).limit(days)

        result = await self.session.execute(query)
        snapshots = list(result.scalars().all())

        data = []
        for snapshot in snapshots:
            for archetype, share in snapshot.archetype_shares.items():
                data.append(
                    {
                        "date": snapshot.snapshot_date.isoformat(),
                        "region": region or "Global",
                        "format": format_type,
                        "archetype": archetype,
                        "share": float(share),
                    }
                )

        return data

    async def _fetch_tournament_results(
        self,
        config: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """Fetch tournament results data."""
        region = config.get("region")
        format_type = config.get("format")
        limit = config.get("limit", 100)

        query = select(Tournament)

        if region:
            query = query.where(Tournament.region == region)
        if format_type:
            query = query.where(Tournament.format == format_type)

        query = query.order_by(Tournament.date.desc()).limit(limit)

        result = await self.session.execute(query)
        tournaments = list(result.scalars().all())

        data = []
        for tournament in tournaments:
            data.append(
                {
                    "tournament_id": tournament.id,
                    "name": tournament.name,
                    "date": tournament.date.isoformat() if tournament.date else None,
                    "region": tournament.region,
                    "format": tournament.format,
                    "tier": tournament.tier,
                    "player_count": tournament.participant_count,
                }
            )

        return data

    async def _fetch_jp_data(
        self,
        config: dict[str, Any],
    ) -> list[dict[str, Any]]:
        """Fetch JP-specific data."""
        days = config.get("days", 30)

        query = (
            select(MetaSnapshot)
            .where(MetaSnapshot.region == "JP")
            .where(MetaSnapshot.best_of == 1)
            .order_by(MetaSnapshot.snapshot_date.desc())
            .limit(days)
        )

        result = await self.session.execute(query)
        snapshots = list(result.scalars().all())

        data = []
        for snapshot in snapshots:
            for archetype, share in snapshot.archetype_shares.items():
                tier_assignments = snapshot.tier_assignments
                data.append(
                    {
                        "date": snapshot.snapshot_date.isoformat(),
                        "archetype": archetype,
                        "share": float(share),
                        "tier": tier_assignments.get(archetype)
                        if tier_assignments
                        else None,
                    }
                )

        return data

    async def _generate_content(
        self,
        data: list[dict[str, Any]],
        format: str,
    ) -> tuple[bytes, list[str]]:
        """Generate formatted content from data."""
        if not data:
            if format == "json":
                return b"[]", []
            elif format == "csv":
                return b"", []
            else:
                xlsx_bytes = await asyncio.to_thread(self._generate_xlsx_sync, [], [])
                return xlsx_bytes, []

        columns = list(data[0].keys())

        if format == "json":
            content = json.dumps(data, indent=2, default=str).encode("utf-8")
            return content, columns

        elif format == "csv":
            output = StringIO()
            writer = csv.DictWriter(output, fieldnames=columns)
            writer.writeheader()
            writer.writerows(data)
            return output.getvalue().encode("utf-8"), columns

        elif format == "xlsx":
            xlsx_content = await asyncio.to_thread(
                self._generate_xlsx_sync, data, columns
            )
            return xlsx_content, columns

        raise ValueError(f"Unknown format: {format}")

    def _generate_xlsx_sync(
        self,
        data: list[dict[str, Any]],
        columns: list[str],
    ) -> bytes:
        """Generate XLSX content synchronously."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()

        # Header row
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Data rows
        for row_idx, row in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name))

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
